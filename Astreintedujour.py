import openpyxl
from datetime import datetime

# Configuration
EXCEL_FILE_PATH = r"\\shares\FTS_Helpdesk\Procédure d'affectation des tickets 2025.xlsx"

def get_astreintes(feuille_astreinte, semaine_actuelle):
    colonne_semaine = None
    for col in range(2, feuille_astreinte.max_column + 1):
        if feuille_astreinte.cell(row=2, column=col).value == f"Semaine {semaine_actuelle}":
            colonne_semaine = col
            break
    
    if colonne_semaine is None:
        print(f"Aucune astreinte trouvée pour la semaine {semaine_actuelle}.")
        return {}
    
    astreintes = {
        "Niveau 1": feuille_astreinte.cell(row=5, column=colonne_semaine).value,
        "Niveau 2": feuille_astreinte.cell(row=6, column=colonne_semaine).value,
        "Cloud": feuille_astreinte.cell(row=7, column=colonne_semaine).value,
        "Network/Security": feuille_astreinte.cell(row=8, column=colonne_semaine).value,
        "SOC": feuille_astreinte.cell(row=9, column=colonne_semaine).value,
    }
    return astreintes

def main():
    try:
        # Ouvrir le fichier Excel
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        
        # Obtenir la semaine actuelle
        semaine_actuelle = datetime.now().isocalendar()[1]
        print(f"Semaine actuelle : {semaine_actuelle}")
        
        # Récupérer les astreintes
        feuille_astreinte = workbook["Astreinte - Congés"]
        astreintes = get_astreintes(feuille_astreinte, semaine_actuelle)
        
        if astreintes:
            print("\nTechniciens d'astreinte pour aujourd'hui:")
            print(f"Niveau 1: {astreintes['Niveau 1']}")
            print(f"Niveau 2: {astreintes['Niveau 2']}")
            print(f"Cloud: {astreintes['Cloud']}")
            print(f"Network/Security: {astreintes['Network/Security']}")
            print(f"SOC: {astreintes['SOC']}")
        else:
            print("Aucune astreinte définie pour cette semaine.")
            
    except Exception as e:
        print(f"Erreur lors de l'ouverture du fichier Excel : {e}")

if __name__ == "__main__":
    main()