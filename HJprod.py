import requests
import json
import openpyxl
import warnings
from datetime import datetime
from collections import defaultdict

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Configuration
base_url = "/api/v3/requests"
auth_token = "xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx"
headers = {"authtoken": auth_token}
verify_ssl = False
processed_ids_file = "processed_requests.txt"
contracts_history_file = "contracts_history.txt"

target_templates = [
    "Sauvegarde", "Stockage", "Replication", 
    "Virtualization", "Serveurs", "Converged Infrastructure", 
    "Switchs", "Autres Assistances"
]

def write_processed_request(request_id, contract_name, sline_value, remaining_days, days_used, total_consumed):
    try:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with open(processed_ids_file, "a") as file:
            file.write(f"{request_id}|{contract_name}|{sline_value}|{remaining_days}|{days_used}|{total_consumed}|{timestamp}\n")
        print(f"Log enregistré pour la requête {request_id} du contrat {contract_name}")
    except Exception as e:
        print(f"Erreur lors de l'écriture dans le fichier des requêtes traitées : {e}")

def read_processed_requests():
    try:
        processed = {}
        with open(processed_ids_file, "r") as file:
            for line in file:
                parts = line.strip().split("|")
                if len(parts) == 7:
                    request_id, contract_name, sline, remaining, used, total_consumed, timestamp = parts
                    processed[request_id] = {
                        'contract_name': contract_name,
                        'sline': sline,
                        'remaining_days': float(remaining),
                        'days_used': float(used),
                        'total_consumed': float(total_consumed),
                        'timestamp': timestamp
                    }
        return processed
    except FileNotFoundError:
        print("Fichier de logs non trouvé, création d'un nouveau fichier.")
        return {}

def load_excel(file_path=r"\\shares\FTS_Helpdesk\Etat assistances Hommes Jours Clients 2025.xlsx", sheet_name="HJ"):
    try:
        print(f"Tentative de chargement du fichier : {file_path}")
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        print(f"Fichier chargé avec succès : {file_path}")
        return sheet, workbook
    except Exception as e:
        print(f"Erreur lors du chargement du fichier Excel : {e}")
        return None, None

def save_excel(workbook, file_path=r"\\shares\FTS_Helpdesk\Etat assistances Hommes Jours Clients 2025.xlsx"):
    try:
        workbook.save(file_path)
        print(f"Fichier Excel '{file_path}' enregistré avec succès.")
    except Exception as e:
        print(f"Erreur lors de l'enregistrement du fichier Excel : {e}")

def get_contract_data(sheet, sline):
    print(f"\nRecherche du SLINE '{sline}' dans le fichier Excel...")
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=3), start=2):
        sline_value = row[0].value
        remaining_days = row[1].value
        consumed_days = row[2].value if row[2].value is not None else 0.0
        
        if sline_value == sline:
            print(f"SLINE trouvé à la ligne {row_idx}!")
            return float(remaining_days) if remaining_days is not None else 0.0, float(consumed_days)
            
    print(f"SLINE '{sline}' non trouvé dans le fichier")
    return None, None

def update_contract_values(sheet, sline, days_used):
    """
    Met à jour à la fois les jours restants (colonne B) et les jours consommés (colonne C)
    """
    for row in sheet.iter_rows(min_row=2, max_col=3):
        if row[0].value == sline:
            # Colonne B: Soustraire les jours utilisés du solde restant
            current_remaining = float(row[1].value) if row[1].value is not None else 0.0
            new_remaining = current_remaining - days_used
            row[1].value = new_remaining

            # Colonne C: Ajouter les jours utilisés au total consommé
            current_consumed = float(row[2].value) if row[2].value is not None else 0.0
            new_total_consumed = current_consumed + days_used
            row[2].value = new_total_consumed

            print(f"\nMise à jour des valeurs pour {sline}:")
            print(f"  Solde initial: {current_remaining}")
            print(f"  Jours utilisés: {days_used}")
            print(f"  Nouveau solde: {new_remaining}")
            print(f"  Total consommé: {new_total_consumed}")

            return new_remaining, new_total_consumed
    return None, None

def process_request(request_id, sheet, workbook, processed_requests):
    if request_id in processed_requests:
        print(f"Requête ID {request_id} déjà traitée, passage.")
        return

    url = f"{base_url}/{request_id}"
    response = requests.get(url, headers=headers, verify=verify_ssl)
    
    if response.status_code == 200:
        request_data = response.json().get("request", {})
        udf_fields = request_data.get("udf_fields", {})
        
        contract_name = request_data.get("subject", "Contrat Inconnu")
        udf_decimal_4232 = udf_fields.get("udf_decimal_4232")  # Jours utilisés
        udf_sline_11101 = udf_fields.get("udf_sline_11101")    # Identifiant ligne

        if udf_decimal_4232 is not None and udf_sline_11101:
            try:
                days_used = float(udf_decimal_4232)
                print(f"\nTraitement de la requête {request_id}:")
                print(f"  Contrat: {contract_name}")
                print(f"  SLINE: {udf_sline_11101}")
                print(f"  Jours utilisés dans cette requête: {days_used}")
                
                # Mettre à jour les valeurs dans Excel
                new_remaining, new_total_consumed = update_contract_values(sheet, udf_sline_11101, days_used)
                
                if new_remaining is not None and new_total_consumed is not None:
                    save_excel(workbook)
                    
                    # Mettre à jour ServiceDesk Plus avec le total des jours consommés
                    input_data = json.dumps({
                        "request": {
                            "udf_fields": {
                                "udf_decimal_4229": new_total_consumed
                            }
                        }
                    })
                    update_response = requests.put(
                        url, 
                        headers=headers, 
                        data={'input_data': input_data}, 
                        verify=verify_ssl
                    )
                    if update_response.status_code == 200:
                        print("Mise à jour ServiceDesk Plus réussie")
                        write_processed_request(
                            request_id,
                            contract_name,
                            udf_sline_11101,
                            new_remaining,
                            days_used,
                            new_total_consumed
                        )
                    else:
                        print(f"Échec de la mise à jour ServiceDesk Plus: {update_response.status_code}")

            except ValueError as e:
                print(f"Erreur de conversion numérique : {e}")
        else:
            print(f"Données manquantes pour la requête {request_id}")
    else:
        print(f"Échec de récupération de la requête {request_id}: {response.status_code}")

def fetch_requests():
    requests_list = []
    start_index = 1
    row_count = 100

    while start_index <= 100:
        input_data = {
            "list_info": {
                "row_count": row_count,
                "start_index": start_index
            }
        }
        try:
            response = requests.get(
                base_url, 
                headers=headers, 
                params={'input_data': json.dumps(input_data)}, 
                verify=verify_ssl
            )
            if response.status_code == 200:
                data = response.json()
                requests_batch = data.get("requests", [])
                if not requests_batch:
                    break
                
                filtered_requests = [
                    req for req in requests_batch 
                    if req.get("template", {}).get("name", "") in target_templates
                ]
                requests_list.extend(filtered_requests)
                
                if not data.get("list_info", {}).get("has_more_rows", False):
                    break
                    
                start_index += row_count
            else:
                print(f"Erreur lors de la récupération des requêtes: {response.status_code}")
                break
        except Exception as e:
            print(f"Erreur lors de la requête HTTP: {e}")
            break

    return requests_list

def main():
    print("\nDémarrage du script de synchronisation...")
    
    sheet, workbook = load_excel()
    if not sheet:
        return

    processed_requests = read_processed_requests()
    requests_list = fetch_requests()

    if not requests_list:
        print("Aucune requête trouvée.")
        return

    print(f"\nTraitement de {len(requests_list)} requêtes...")
    
    for request in requests_list:
        template_name = request.get("template", {}).get("name", "")
        request_id = request.get("id")
        if template_name in target_templates and request_id:
            print(f"\nTraitement de la requête ID {request_id} avec le template '{template_name}'...")
            process_request(request_id, sheet, workbook, processed_requests)

    print("\nSynchronisation terminée.")

if __name__ == "__main__":
    main()
