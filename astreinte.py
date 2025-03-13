import openpyxl
from datetime import datetime

# Configuration
EXCEL_FILE_PATH = r"Procédure d'affectation des tickets 2025.xlsx"

def get_astreinte_niveau2(date_demandee):
    """
    Récupère l'astreinte niveau 2 pour une date donnée
    
    Args:
        date_demandee (str): Date au format 'DD/MM/YYYY'
    """
    try:
        # Convertir la date string en objet datetime
        date_obj = datetime.strptime(date_demandee, '%d/%m/%Y')
        
        # Obtenir le numéro de la semaine
        semaine = date_obj.isocalendar()[1]
        
        # Ouvrir le fichier Excel
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        feuille_astreinte = workbook["Astreinte - Congés"]
        
        # Chercher la colonne correspondant à la semaine
        colonne_semaine = None
        for col in range(2, feuille_astreinte.max_column + 1):
            if feuille_astreinte.cell(row=2, column=col).value == f"Semaine {semaine}":
                colonne_semaine = col
                break
        
        if colonne_semaine is None:
            return f"Aucune astreinte trouvée pour la semaine {semaine}"
        
        # Récupérer le nom de l'astreinte niveau 2
        astreinte_niveau2 = feuille_astreinte.cell(row=6, column=colonne_semaine).value
        
        return {
            "date": date_demandee,
            "semaine": semaine,
            "astreinte_niveau2": astreinte_niveau2
        }
        
    except ValueError:
        return "Erreur: Format de date incorrect. Utilisez le format DD/MM/YYYY"
    except Exception as e:
        return f"Erreur lors de la lecture du fichier Excel: {str(e)}"

def main():
    # Demander la date à l'utilisateur
    date_demandee = input("Entrez la date (format DD/MM/YYYY): ")
    
    # Obtenir et afficher l'astreinte
    resultat = get_astreinte_niveau2(date_demandee)
    
    if isinstance(resultat, dict):
        print(f"\nRésultats pour le {resultat['date']}:")
        print(f"Semaine: {resultat['semaine']}")
        print(f"Astreinte Niveau 2: {resultat['astreinte_niveau2']}")
    else:
        print(resultat)

if __name__ == "__main__":
    main()
