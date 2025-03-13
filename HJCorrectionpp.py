import requests
import json
import openpyxl
import warnings
from datetime import datetime
import time
from collections import defaultdict

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Configuration
base_url = "/api/v3/requests"
auth_token = "xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx"
headers = {"authtoken": auth_token}
verify_ssl = False
processed_requests_file = "processed_requests.txt"

def log_debug(message):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {message}")

def load_excel(file_path=r"\\shares\FTS_Helpdesk\Etat assistances Hommes Jours Clients 2025.xlsx", sheet_name="HJ"):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet, workbook
    except Exception as e:
        log_debug(f"Erreur lors du chargement du fichier Excel : {e}")
        return None, None

def find_contract_row(sheet, contract_name):
    """Trouve la ligne correspondant au contrat dans Excel"""
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == contract_name:
            return row
    return None

def load_processed_requests():
    """Charge et organise l'historique des tickets par SLINE et nom de contrat"""
    contracts = defaultdict(lambda: defaultdict(dict))
    try:
        with open(processed_requests_file, "r") as file:
            for line in file:
                parts = line.strip().split("|")
                if len(parts) == 7:
                    request_id, sline, contract_name, remaining, days_used, total_consumed, timestamp = parts
                    contracts[sline][contract_name][request_id] = {
                        'remaining_days': float(remaining),
                        'days_used': float(days_used),
                        'total_consumed': float(total_consumed),
                        'timestamp': timestamp
                    }
    except FileNotFoundError:
        log_debug("Fichier processed_requests.txt non trouvé")
    return contracts

def update_processed_requests(sline, request_id, contract_name, remaining_days, new_days_used, new_total_consumed):
    """Met à jour l'historique d'un ticket dans processed_requests.txt"""
    lines = []
    
    try:
        with open(processed_requests_file, "r") as file:
            lines = [line for line in file if line.strip().split("|")[0] != request_id]
    except FileNotFoundError:
        log_debug("Fichier processed_requests.txt non trouvé")
        return False

    # Ajouter la nouvelle ligne
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    new_line = f"{request_id}|{sline}|{contract_name}|{remaining_days}|{new_days_used}|{new_total_consumed}|{timestamp}\n"
    lines.append(new_line)

    # Écrire toutes les lignes dans le fichier
    with open(processed_requests_file, "w") as file:
        file.writelines(lines)
    
    log_debug(f"Historique mis à jour: {new_line.strip()}")
    return True

def update_excel(sheet, workbook, contract_name, new_days_used, old_days_used):
    """Met à jour le Nouveau Solde dans la colonne B pour le contrat spécifié"""
    contract_row = find_contract_row(sheet, contract_name)
    
    if contract_row is None:
        log_debug(f"Contrat {contract_name} non trouvé dans Excel")
        return False, None
        
    try:
        # Lire le solde actuel dans la colonne B
        old_remaining = float(sheet.cell(row=contract_row, column=2).value or 0)
        
        # Calculer le nouveau solde
        new_remaining = old_remaining + old_days_used - new_days_used
        
        # Mettre à jour la colonne B (Nouveau Solde)
        sheet.cell(row=contract_row, column=2).value = new_remaining
        
        try:
            workbook.save(r"\\shares\FTS_Helpdesk\Etat assistances Hommes Jours Clients 2025.xlsx")
            log_debug(f"Excel mis à jour pour le contrat {contract_name} - Nouveau solde: {new_remaining}")
            return True, new_remaining
        except Exception as e:
            log_debug(f"Erreur lors de la sauvegarde Excel: {e}")
            return False, None
            
    except Exception as e:
        log_debug(f"Erreur lors de la mise à jour Excel pour le contrat {contract_name}: {e}")
        return False, None

def check_ticket_updates():
    log_debug("\nDémarrage de la vérification...")
    
    sheet, workbook = load_excel()
    if not sheet:
        return

    contracts = load_processed_requests()
    
    for sline, contract_data in contracts.items():
        log_debug(f"\nVérification du contrat {sline}")
        
        for contract_name, tickets in contract_data.items():
            for request_id, old_data in tickets.items():
                url = f"{base_url}/{request_id}"
                try:
                    response = requests.get(url, headers=headers, verify=verify_ssl)
                    if response.status_code == 200:
                        request_data = response.json().get("request", {})
                        udf_fields = request_data.get("udf_fields", {})
                        
                        try:
                            current_days_used = float(udf_fields.get("udf_decimal_4232", 0))
                        except (TypeError, ValueError):
                            current_days_used = 0
                        
                        old_days_used = old_data['days_used']
                        
                        # Vérifier si le nombre de jours consommés a changé
                        if abs(current_days_used - old_days_used) > 0.001:
                            log_debug(f"\nChangement détecté pour ticket {request_id}:")
                            log_debug(f"  Ancien HJ: {old_days_used}")
                            log_debug(f"  Nouveau HJ: {current_days_used}")
                            
                            # Mettre à jour Excel
                            success, new_remaining = update_excel(
                                sheet, workbook,
                                contract_name,
                                current_days_used, 
                                old_days_used
                            )
                            
                            if success:
                                # Mettre à jour l'historique dans le fichier .txt
                                update_processed_requests(
                                    sline, request_id,
                                    contract_name,
                                    new_remaining,
                                    current_days_used,
                                    old_data['total_consumed']  # Garder l'ancienne valeur du total
                                )
                                log_debug(f"Mise à jour complète réussie pour le ticket {request_id}")
                        
                except Exception as e:
                    log_debug(f"Erreur vérification ticket {request_id}: {e}")

def main():
    while True:
        try:
            check_ticket_updates()
            log_debug("\nAttente de 10 secondes avant prochaine vérification...")
            time.sleep(10)
        except Exception as e:
            log_debug(f"Erreur dans la boucle principale: {e}")
            time.sleep(60)

if __name__ == "__main__":
    main()
