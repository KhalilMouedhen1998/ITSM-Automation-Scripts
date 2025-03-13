import requests
import pandas as pd
from datetime import datetime
import openpyxl
import subprocess
import re

# Configuration
API_URL = "https://10.10.30.75/api/v3/requests"
API_TOKEN = "0C865686-02A7-45EE-B13A-8C22A1066429"
EXCEL_FILE_PATH = r"Procédure d'affectation des tickets 2025.xlsx"

def get_tech_with_least_tickets():
    """Exécute le script tech-niv1-infra.py et récupère le nom du technicien"""
    try:
        result = subprocess.run(['python', 'tech-niv1-infra.py'], capture_output=True, text=True)
        output = result.stdout
        # Utiliser une expression régulière pour extraire le nom du technicien
        match = re.search(r"Le technicien (.*?) a le moins de tickets", output)
        if match:
            return match.group(1)
        return None
    except Exception as e:
        print(f"Erreur lors de l'exécution de tech-niv1-infra.py : {e}")
        return None

# Fonction pour récupérer la date de création du ticket via l'API
def get_ticket_created_time():
    headers = {"authtoken": API_TOKEN}
    input_data = '''{
        "list_info": {
            "row_count": 1,
            "start_index": 1
        }
    }'''
    params = {'input_data': input_data}
    response = requests.get(API_URL, headers=headers, params=params, verify=False)
    
    if response.status_code == 200:
        data = response.json()
        created_time = data["requests"][0]["created_time"]["display_value"]
        return datetime.strptime(created_time, "%d/%m/%Y %I:%M %p"), data["requests"][0]["id"]
    else:
        print(f"Erreur API : {response.status_code}")
        return None, None

# Fonction pour récupérer les astreintes
def get_astreintes(feuille_astreinte, semaine_actuelle):
    colonne_semaine = None
    for col in range(2, feuille_astreinte.max_column + 1):
        if feuille_astreinte.cell(row=2, column=col).value == f"Semaine {semaine_actuelle}":
            colonne_semaine = col
            break
    
    if colonne_semaine is None:
        print(f"Aucune astreinte trouvée pour la semaine {semaine_actuelle}.")
        return {}
    
    astreintes = {
        "Niveau 1": feuille_astreinte.cell(row=5, column=colonne_semaine).value,
        "Niveau 2": feuille_astreinte.cell(row=6, column=colonne_semaine).value,
        "Cloud": feuille_astreinte.cell(row=7, column=colonne_semaine).value,
        "Network/Security": feuille_astreinte.cell(row=8, column=colonne_semaine).value,
        "SOC": feuille_astreinte.cell(row=9, column=colonne_semaine).value,
    }
    return astreintes

# Fonction pour envoyer une requête API
def send_api_request(request_id, technician_name, level):
    technician_name = technician_name.strip()  # Retirer les espaces inutiles
    url = f"{API_URL}/{request_id}"
    headers = {"authtoken": API_TOKEN}
    input_data = f'''{{
        "request": {{
            "technician": {{ "name": "{technician_name}" }},
            "level": {{ "name": "{level}" }}
        }}
    }}'''

    print(f"Data envoyée à l'API : {input_data}")
    
    data = {'input_data': input_data}
    response = requests.put(url, headers=headers, data=data, verify=False)
    if response.status_code == 200:
        print("Technicien affecté avec succès")
    else:
        print(f"Erreur dans l'affectation : {response.text}")

# Fonction principale
def main():
    created_datetime, request_id = get_ticket_created_time()
    if not created_datetime or not request_id:
        return
    
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    except Exception as e:
        print(f"Erreur lors de l'ouverture du fichier Excel : {e}")
        return
    
    semaine_actuelle = created_datetime.isocalendar()[1]
    print(f"Semaine actuelle : {semaine_actuelle}")
    
    heure_creation = created_datetime.time()
    if heure_creation >= datetime.strptime("17:30", "%H:%M").time() or heure_creation < datetime.strptime("08:29", "%H:%M").time():
        print("Le ticket est en astreinte.")
        feuille_astreinte = workbook["Astreinte - Congés"]
        astreintes = get_astreintes(feuille_astreinte, semaine_actuelle)
        
        if astreintes:
            tech_niv1 = astreintes["Niveau 1"]
            tech_niv2 = astreintes["Niveau 2"]
            tech_cloud = astreintes["Cloud"]
            tech_network_security = astreintes["Network/Security"]
            tech_soc = astreintes["SOC"]

            print(f"Astreinte Niveau 1 : {tech_niv1}")
            print(f"Astreinte Niveau 2 : {tech_niv2}")
            print(f"Astreinte Cloud : {tech_cloud}")
            print(f"Astreinte Network/Security : {tech_network_security}")
            print(f"Astreinte SOC : {tech_soc}")
            
            tech_to_assign = tech_niv1
            level_to_assign = "Tier 1"
            send_api_request(request_id, tech_to_assign, level_to_assign)
        else:
            print("Aucune astreinte définie pour cette semaine.")
    else:
        print("Le ticket n'est pas en astreinte. Affectation normale.")
        # Récupérer le technicien avec le moins de tickets
        tech_to_assign = get_tech_with_least_tickets()
        if tech_to_assign:
            url = f"{API_URL}/{request_id}"
            headers = {"authtoken": API_TOKEN}
            input_data = f'''{{
                "request": {{
                    "group": {{
                        "name": "Service desk"
                    }},
                    "level": {{
                        "name": "Tier 1"
                    }},
                    "technician": {{
                        "name": "{tech_to_assign}"
                    }}
                }}
            }}'''
            data = {'input_data': input_data}
            response = requests.put(url, headers=headers, data=data, verify=False)
            print(response.text)
        else:
            print("Erreur : Impossible de déterminer le technicien à affecter")

if __name__ == "__main__":
    main()