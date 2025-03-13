import requests
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# API de base
base_url = "/api/v3/requests"
headers = {"authtoken": ""}

# Liste des templates ciblés
target_templates = [
    "Sauvegarde", "Stockage", "Replication", 
    "Virtualization", "Serveurs", "Converged Infrastructure", 
    "Switchs", "Autres Assistances"
]

# Fonction pour récupérer une page spécifique
def fetch_page(start_index):
    input_data = {
        "list_info": {
            "row_count": 100,
            "start_index": start_index,
        }
    }
    params = {'input_data': json.dumps(input_data)}
    response = requests.get(base_url, headers=headers, params=params, verify=False)
    return response.json()

# Fonction pour récupérer les détails d'une demande
def fetch_request_details(request_id):
    url = f"{base_url}/{request_id}"
    response = requests.get(url, headers=headers, verify=False)
    return response.json()

# Récupération des pages et extraction des IDs
all_request_ids = []
for start_index in [1, 101, 201]:  # Pages 1, 2 et 3
    page_data = fetch_page(start_index)
    if "requests" in page_data:
        all_request_ids.extend([req["id"] for req in page_data["requests"]])

# Structure pour stocker les données par compte
accounts_data = {}

# Récupération des détails pour chaque ID
for request_id in all_request_ids:
    details = fetch_request_details(request_id)
    if details.get("response_status", {}).get("status_code") == 2000:
        request = details["request"]
        template_name = request["template"]["name"]

        # Filtrer par template
        if template_name in target_templates:
            account_name = request["account"]["name"]

            # Extraire les informations nécessaires
            ticket_data = {
                "Ticket ID": request["id"],
                "Creation Date": request["created_time"]["display_value"],
                "H/J Consommés": request["udf_fields"].get("udf_decimal_4232", "N/A"),
                "Reste H/J": request["udf_fields"].get("udf_decimal_4233", "N/A"),
                "Asset": request["primary_asset"]["name"] if request.get("primary_asset") else "N/A",
                "Subject": request["subject"],
                "Actions": ", ".join([str(task["id"]) for task in request.get("request_template_task_ids", [])]),
                "Technician": request["technician"]["name"] if request.get("technician") else "N/A",
                "Requester": request["requester"]["name"]
            }

            # Ajouter les données au compte correspondant
            if account_name not in accounts_data:
                accounts_data[account_name] = []
            accounts_data[account_name].append(ticket_data)

# Exportation vers Excel
wb = Workbook()
for account, tickets in accounts_data.items():
    # Créer une feuille pour chaque compte
    ws = wb.create_sheet(title=account[:31])  # Limité à 31 caractères
    ws.append(["Ticket ID", "Creation Date", "H/J Consommés", "Reste H/J",
               "Asset", "Subject", "Actions", "Technician", "Requester"])

    # Ajouter les données
    for ticket in tickets:
        ws.append([ticket["Ticket ID"], ticket["Creation Date"], ticket["H/J Consommés"],
                   ticket["Reste H/J"], ticket["Asset"], ticket["Subject"],
                   ticket["Actions"], ticket["Technician"], ticket["Requester"]])

    # Ajuster la largeur des colonnes à 30mm (~ 85 pixels)
    for col in range(1, 10):  # 1 à 9 colonnes
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 30  # Largeur définie en caractères (environ 30mm)

# Supprimer la feuille par défaut "Sheet" si elle existe
if "Sheet" in wb.sheetnames:
    wb.remove(wb["Sheet"])

# Sauvegarder le fichier Excel
wb.save("Filtered_Requests_by_Account.xlsx")
print("Fichier Excel généré : Filtered_Requests_by_Account.xlsx")
